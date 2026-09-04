# -*- coding: utf-8 -*-
"""Объединённый реестр каналов продажи элементов питания в РФ.

Источники:
  A) мой поиск по рынку (реестр из build_registry.py, 103 площадки);
  B) партнёрский список заказчика с ИНН и сайтами (415 строк, 374 домена).
Пересечение — 34 домена. Файл сводит оба среза в один мониторинг.
"""
import json, re, collections, importlib.util

FONT = "Arial"
ROOT = "/home/user/aitim"

spec = importlib.util.spec_from_file_location("b", f"{ROOT}/build_registry.py")
REG = importlib.util.module_from_spec(spec); spec.loader.exec_module(REG)

CLASSIFIER = REG.CLASSIFIER
CLS_NAME = REG.CLS_NAME


def dom(u):
    if not u:
        return ""
    u = str(u).strip().lower()
    if u in ("нет", "none", "-", "—", "", "интернет-магазин", "ооо", "www.", "narod.ru", "spb.ru",
             "ru.com", "vk.com", "wixsite.com", "tomsk.ru", "modi", "regstaer"):
        return ""
    m = re.match(r"^(?:https?://)?(?:www\.)?([^/\s?]+)", u)
    d = m.group(1) if m else ""
    return d if "." in d and " " not in d else ""


def base(d):
    p = d.split(".")
    if len(p) >= 3 and p[-2] in ("co", "com", "net", "org"):
        return ".".join(p[-3:])
    return ".".join(p[-2:]) if len(p) >= 2 else d


# ---- ручные соответствия домен -> (направление, подкатегория) ----
KNOWN = {
 # специализированные по элементам питания
 "netizbattery.ru": ("SPEC", "Элементы питания"), "spb-element.ru": ("SPEC", "Элементы питания"),
 "2b2batkomplekt.ru": ("SPEC", "Элементы питания"), "batalux.ru": ("SPEC", "Элементы питания"),
 "perfeo.ru": ("SPEC", "Производитель элементов питания и электроники"),
 "energnew.ru": ("SPEC", "Элементы питания и энергетика"),
 "alpha-energy.ru": ("SPEC", "Элементы питания и энергетика"),
 "energia24.ru": ("SPEC", "Элементы питания и энергетика"),
 "energomix.ru": ("SPEC", "Элементы питания и энергетика"),
 "energosfera.ru": ("SPEC", "Элементы питания и энергетика"),
 "u-energo.ru": ("SPEC", "Элементы питания и энергетика"),
 # аптеки и медтехника
 "apteka.ru": ("NICHE", "Аптечный ритейл"), "aptekus.ru": ("NICHE", "Аптечный ритейл"),
 "autopteka.ru": ("NICHE", "Аптечный ритейл"), "zdravcity.ru": ("NICHE", "Аптечный ритейл"),
 "pharmeconom.ru": ("NICHE", "Аптечный ритейл"), "protek.ru": ("NICHE", "Фарм-дистрибуция"),
 "cardiomed.su": ("NICHE", "Медтехника"), "myocard.ru": ("NICHE", "Медтехника"),
 "spektr-med.com": ("NICHE", "Медтехника"), "sphinx-md.ru": ("NICHE", "Медтехника"),
 # охранные системы
 "bolid.ru": ("NICHE", "Охранные системы"), "rubezh.ru": ("NICHE", "Охранные системы"),
 "rielta.ru": ("NICHE", "Охранные системы"), "teko.biz": ("NICHE", "Охранные системы"),
 "securon.ru": ("NICHE", "Охранные системы"), "sct.ru": ("NICHE", "Охранные системы"),
 "system-r.ru": ("NICHE", "Охранные системы"),
 # авто
 "euroauto.ru": ("NICHE", "Автозапчасти"), "bi-bi.ru": ("NICHE", "Автотовары"),
 "lecar.ru": ("NICHE", "Автотовары"), "v-avto.ru": ("NICHE", "Автотовары"),
 "forwardauto.ru": ("NICHE", "Автотовары"), "avtodt.ru": ("NICHE", "Автотовары"),
 "lal-auto.ru": ("NICHE", "Автотовары"), "dakomavt.ru": ("NICHE", "Автотовары"),
 "automech.su": ("NICHE", "Автотовары"), "зарулем-шоп.рф": ("NICHE", "Автотовары"),
 "avtomarket.com": ("NICHE", "Автотовары"), "ав-торг.рф": ("NICHE", "Автотовары"),
 "autofon.ru": ("NICHE", "GPS-трекеры и мониторинг"), "navitel.ru": ("NICHE", "Навигация"),
 "tricolor.ru": ("NICHE", "ТВ-оборудование"),
 # АЗС и топливо
 "neftm.ru": ("QCOM", "АЗС-ритейл"), "tranzit-oil.com": ("QCOM", "АЗС-ритейл"),
 "oilcity.ru": ("QCOM", "АЗС-ритейл"), "tdrusoil.ru": ("QCOM", "АЗС-ритейл"),
 "dme-dutyfree.ru": ("QCOM", "Travel retail / duty free"),
 "sbermarket.ru": ("QCOM", "Агрегатор доставки (ныне Купер)"),
 "tkvprok.ru": ("QCOM", "Онлайн-гипермаркет"),
 # фото, музыка, хобби, спорт, детские
 "konturfoto.ru": ("NICHE", "Фототовары"), "foto-sivma.ru": ("NICHE", "Фототовары"),
 "foto4u.su": ("NICHE", "Фототовары"), "muztorg.ru": ("NICHE", "Музыкальные инструменты"),
 "nokturn.ru": ("NICHE", "Музыкальные инструменты"), "e-nokturn.ru": ("NICHE", "Музыкальные инструменты"),
 "leonardo.ru": ("NICHE", "Хобби и творчество"), "souzmultpark.ru": ("NICHE", "Детские товары"),
 "ural.toys": ("NICHE", "Детские товары"), "7mya.ru": ("NICHE", "Детские товары"),
 "igla.ru": ("NICHE", "Рукоделие и хобби"), "assortyidey.ru": ("NICHE", "Хобби и творчество"),
 "huntworld.ru": ("NICHE", "Охота и туризм"), "stickhunt.ru": ("NICHE", "Охота и туризм"),
 "zenitco.ru": ("NICHE", "Оружейные аксессуары"), "rainblade.pro": ("NICHE", "Охота и туризм"),
 "smoker.su": ("NICHE", "Табак и сопутствующие"), "eroticfantasy.ru": ("NICHE", "Товары для взрослых"),
 "sibsemena.ru": ("NICHE", "Сад и огород"), "sadhoztorg.ru": ("NICHE", "Сад и хозтовары"),
 # парфюмерия и дрогери
 "goldapple.ru": ("DROG", "Парфюмерия и косметика"), "podrygka.ru": ("DROG", "Дрогери"),
 "nevaparfum.ru": ("DROG", "Парфюмерия"), "parfum-lider.ru": ("DROG", "Парфюмерия"),
 "parfum-bhs.ru": ("DROG", "Парфюмерия"), "interraparfum.ru": ("DROG", "Парфюмерия"),
 "scent.ru": ("DROG", "Парфюмерия"), "marafett.ru": ("DROG", "Дрогери"),
 "lamel.shop": ("DROG", "Косметика"), "letto.ru": ("DROG", "Дрогери"),
 # продуктовый ритейл
 "maria-ra.ru": ("FMCG", "Продуктовые сети"), "slata.ru": ("FMCG", "Продуктовые сети"),
 "slata.com": ("FMCG", "Продуктовые сети"), "ярче.рф": ("FMCG", "Продуктовые сети"),
 "miratorg.ru": ("FMCG", "Продуктовые сети"), "novex.ru": ("FMCG", "Дрогери и продукты"),
 "maxi-retail.ru": ("FMCG", "Продуктовые сети"), "avoska.ru": ("FMCG", "Продуктовые сети"),
 "dobrotsen.ru": ("HARD", "Жёсткий дискаунтер"), "prostor-market.ru": ("HARD", "Дискаунтер"),
 "tgabsolut-shop.ru": ("FMCG", "Продуктовые сети"), "narodny38.ru": ("FMCG", "Продуктовые сети"),
 "nahodka-magazin.ru": ("FMCG", "Продуктовые сети"), "kirmarket.ru": ("FMCG", "Продуктовые сети"),
 "irkmarket.ru": ("FMCG", "Продуктовые сети"), "krasyar.ru": ("FMCG", "Продуктовые сети"),
 "ts-7dney.ru": ("FMCG", "Продуктовые сети"), "sparural.ru": ("FMCG", "Продуктовые сети"),
 "amwine.ru": ("FMCG", "Алкомаркеты"), "pivko24.ru": ("FMCG", "Алкомаркеты"),
 "vinograd.shop": ("FMCG", "Алкомаркеты"), "русский-разгуляйка.рф": ("FMCG", "Продуктовые сети"),
 "lifemart.ru": ("FMCG", "Продуктовые сети"), "famil.ru": ("HARD", "Офпрайс"),
 "offprice.eu": ("HARD", "Офпрайс"), "leomax.ru": ("HARD", "ТВ-ритейл"),
 "магазин-рубль.рф": ("HARD", "Фикс-прайс"), "tovarbezpereplat.ru": ("HARD", "Дискаунтер"),
 # DIY, дом, хозтовары
 "poryadok.ru": ("DIY", "DIY и товары для дома"), "tvoydom.ru": ("DIY", "Товары для дома"),
 "oma.by": ("DIY", "DIY-гипермаркеты"), "krephaus.ru": ("DIY", "Крепёж и инструмент"),
 "vintboltovich.ru": ("DIY", "Крепёж и инструмент"), "добрострой.рф": ("DIY", "Стройматериалы"),
 "ekonomstroy.ru": ("DIY", "Стройматериалы"), "eldvor.ru": ("DIY", "Электрика и стройматериалы"),
 "gipermarketdom.ru": ("DIY", "Товары для дома"), "vashdom24.ru": ("DIY", "Товары для дома"),
 "suntehhoztorg.ru": ("DIY", "Сантехника и хозтовары"), "hozsfera.ru": ("DIY", "Хозтовары"),
 "hozsfera24.ru": ("DIY", "Хозтовары"), "hoz-rb.ru": ("DIY", "Хозтовары"),
 "hozmarket.store": ("DIY", "Хозтовары"), "opthz.ru": ("DIY", "Хозтовары"),
 "poiskhome.ru": ("DIY", "Товары для дома"),
 # электроника и техника
 "xcom-shop.ru": ("EL", "Компьютерная техника"), "netlab.ru": ("EL", "IT-дистрибуция"),
 "ret.ru": ("EL", "Ритейл техники"), "magazin-elektronika.ru": ("EL", "Электроника"),
 "elektronika126.ru": ("EL", "Электроника"), "vsesmart.ru": ("EL", "Электроника"),
 "smart-shop.pro": ("EL", "Электроника"), "digitalpapa.ru": ("EL", "Электроника"),
 "radiomir96.ru": ("EL", "Радиодетали"), "radiomirnsk.ru": ("EL", "Радиодетали"),
 "megafon.ru": ("EL", "Салоны связи"), "moscow.megafon.ru": ("EL", "Салоны связи"),
 "tdelekom.ru": ("EL", "Телеком-оборудование"), "planets.ru": ("EL", "Электроника"),
 "avatonshop.ru": ("EL", "Электроника"), "mirdiodov.info": ("EL", "Светотехника"),
 "impuls-sveta.ru": ("EL", "Светотехника"), "apollo-lux.ru": ("EL", "Светотехника"),
 "tesla39.ru": ("EL", "Электроника и электрика"),
 # b2b, опт, канцелярия
 "informat.ru": ("B2B", "Канцтовары опт"), "elimkanz.ru": ("B2B", "Канцтовары"),
 "giftspromo.ru": ("B2B", "Промо и сувениры"), "pg.pro": ("B2B", "Промо и сувениры"),
 "rpk.group": ("B2B", "Промо и сувениры"), "megapack-russia.ru": ("B2B", "Упаковка"),
 "союзупак.рф": ("B2B", "Упаковка"), "mr-pak.ru": ("B2B", "Упаковка"),
 "retailhoreca.ru": ("B2B", "HoReCa-снабжение"), "optomhim.ru": ("B2B", "Бытовая химия опт"),
 "mygrass.ru": ("B2B", "Бытовая химия"), "himtver.ru": ("B2B", "Бытовая химия"),
 "poroshki100.ru": ("B2B", "Бытовая химия"),
 "elektro.ru": ("B2B", "Электротехническая дистрибуция"),
 "eltsnab.ru": ("B2B", "Электротехническая дистрибуция"),
 "electro-mpo.ru": ("B2B", "Электротехническая дистрибуция"),
 "electro-master.ru": ("B2B", "Электротехническая дистрибуция"),
 "electrica-penza.ru": ("B2B", "Электротехническая дистрибуция"),
 "electro58.ru": ("B2B", "Электротехническая дистрибуция"),
 "krdelectro.ru": ("B2B", "Электротехническая дистрибуция"),
 "elkomp.ru": ("B2B", "Электротехническая дистрибуция"),
 "el-on.ru": ("B2B", "Электротехническая дистрибуция"),
 "elteam.pro": ("B2B", "Электротехническая дистрибуция"),
 "minimaks.ru": ("B2B", "Электротехническая дистрибуция"),
 "voltacom.ru": ("B2B", "Электротехническая дистрибуция"),
 "rubilnik.ru": ("B2B", "Электротехническая дистрибуция"),
 "эльсити.рф": ("B2B", "Электротехническая дистрибуция"),
 "ars-elektra.ru": ("B2B", "Электротехническая дистрибуция"),
 "progress-etk.ru": ("B2B", "Электротехническая дистрибуция"),
 "c-e.ru": ("B2B", "Электронные компоненты"),
 "unit2000.ru": ("B2B", "Электронные компоненты"),
 "texkom.ru": ("B2B", "Электронные компоненты"),
}


KNOWN.update({
 "relef.ru": ("B2B", "Опт канцтоваров и хозтоваров"),
 "serviko.ru": ("B2B", "Мультикатегорийная дистрибуция"),
 "swnn.ru": ("B2B", "FMCG-дистрибуция"),
 "tdagat.ru": ("B2B", "Мультикатегорийная дистрибуция"),
 "tdagat-shop.ru": ("B2B", "Мультикатегорийная дистрибуция"),
 "tdmega.com": ("B2B", "FMCG-дистрибуция"),
 "raznotorg.com": ("B2B", "Мультикатегорийный опт"),
 "sklad-ufa.com": ("B2B", "Мультикатегорийный опт"),
 "kopilkago.ru": ("FMCG", "Продуктовые сети"),
 "drujba-tc.ru": ("FMCG", "Продуктовые сети"),
 "victoria-group.ru": ("FMCG", "Продуктовые сети"),
 "shamsa.net": ("FMCG", "Продуктовые сети"),
 "b54.ru": ("FMCG", "Продуктовые сети"),
 "magnum.kz": ("FMCG", "Продуктовые сети (Казахстан)"),
 "globus-online.kg": ("FMCG", "Продуктовые сети (Киргизия)"),
 "umaigroup.kg": ("FMCG", "Продуктовые сети (Киргизия)"),
 "azmak.kz": ("B2B", "Дистрибуция (Казахстан)"),
 "rustel.kg": ("B2B", "Дистрибуция (Киргизия)"),
 "oma.by": ("DIY", "DIY-гипермаркеты (Беларусь)"),
 "sila.by": ("DIY", "DIY (Беларусь)"),
 "azuma.store": ("DROG", "Дрогери"),
 "plenki.net": ("NICHE", "Плёнки и аксессуары"),
 "revo-udc.com": ("NICHE", "Зарядная инфраструктура"),
 "pasker.ru": ("NICHE", "Автотовары"),
 "avtopasker.ru": ("NICHE", "Автотовары"),
 "magistral-nn.ru": ("NICHE", "Автотовары"),
 "termopuls.ru": ("B2B", "Инженерное оборудование"),
 "pk-tp.ru": ("B2B", "Промышленное снабжение"),
 "gts114.ru": ("B2B", "Снабжение"),
 "konnekt58.ru": ("EL", "Электроника и связь"),
 "di-house.ru": ("EL", "IT-дистрибуция"),
 "dors.com": ("EL", "Банковское оборудование"),
 "csat.ru": ("NICHE", "Охранные системы"),
 "aka-scan.ru": ("NICHE", "Приборы и детекторы"),
 "s-globus.ru": ("B2B", "Мультикатегорийный опт"),
 "vertical.ru": ("B2B", "Мультикатегорийный опт"),
 "cellfaktor.ru": ("SPEC", "Элементы питания"),
 "neptunsale.ru": ("B2B", "Мультикатегорийный опт"),
 "usta.group": ("B2B", "Мультикатегорийный опт"),
 "yurinat.ru": ("B2B", "Мультикатегорийный опт"),
})

CIS_TLD = (".kz", ".kg", ".by", ".uz", ".md", ".am")


def is_cis(d):
    return d.endswith(CIS_TLD)


# ---- эвристика по названию партнёра и домену ----
HEUR = [
 (r"БАТАРЕ|БЭТТЕР|ЭЛЕМЕНТ ПИТ|BATTER|BATT|ЭЛЕМЕНТ ООО", ("SPEC", "Элементы питания")),
 (r"ЭНЕРГ|ENERG|АККУМ", ("SPEC", "Элементы питания и энергетика")),
 (r"АПТЕК|ФАРМ|PHARM|ZDRAV|МЕД|MED|КАРДИО", ("NICHE", "Аптеки и медтехника")),
 (r"ОХРАН|БЕЗОПАСН|SECUR|РУБЕЖ|БОЛИД", ("NICHE", "Охранные системы")),
 (r"АВТО|AUTO|АВТ|ШИН|ЗАПЧАСТ", ("NICHE", "Автотовары")),
 (r"ФОТО|FOTO|PHOTO", ("NICHE", "Фототовары")),
 (r"МУЗЫК|MUZ|НОКТЮРН", ("NICHE", "Музыкальные товары")),
 (r"ОХОТ|HUNT|РЫБОЛ|ТУРИЗМ|ОРУЖ|ARMS", ("NICHE", "Охота, оружие и туризм")),
 (r"ЗОО|ZOO|PET", ("NICHE", "Зоотовары")),
 (r"ДЕТ|TOYS|ИГРУШ|KIDS", ("NICHE", "Детские товары")),
 (r"ПАРФЮМ|PARFUM|КОСМЕТ|SCENT|БЬЮТИ|BEAUTY", ("DROG", "Парфюмерия и косметика")),
 (r"АЗС|НЕФТ|OIL|ТОПЛИВ|ЗАПРАВ", ("QCOM", "АЗС-ритейл")),
 (r"ЭЛЕКТР|ELECTR|ELEKTR|СВЕТ|LIGHT|ВОЛЬТ|VOLT|ЭТК|КАБЕЛЬ", ("B2B", "Электротехническая дистрибуция")),
 (r"КОМПОНЕНТ|РАДИО|RADIO|ЧИП|CHIP", ("B2B", "Электронные компоненты")),
 (r"КАНЦ|ОФИС|OFFICE|БУМАГ|INFORMAT|СЕКРЕТАР", ("B2B", "Канцтовары и офис")),
 (r"УПАКОВ|PACK|ПАК|ТАРА", ("B2B", "Упаковка")),
 (r"ХИМ|CHEM|МОЮЩ|ПОРОШ", ("B2B", "Бытовая химия")),
 (r"ХОЗ|HOZ|САНТЕХ|ПОСУД", ("DIY", "Хозтовары и товары для дома")),
 (r"СТРОЙ|STROY|РЕМОНТ|КРЕПЁЖ|КРЕПЕЖ|ИНСТРУМЕНТ|DIY|ДОМ|DOM", ("DIY", "DIY и стройматериалы")),
 (r"МАРКЕТ|MARKET|СУПЕРМАРКЕТ|ПРОДУКТ|ГАСТРОНОМ|ПРОДМИР|ТОРГСЕРВИС", ("FMCG", "Продуктовая розница")),
 (r"АЛКО|ВИН|WINE|ПИВ|BEER", ("FMCG", "Алкомаркеты")),
 (r"ДИСКАУНТ|ФИКС|РУБЛ|ЭКОНОМ|ДОБРОЦЕН", ("HARD", "Дискаунтеры")),
 (r"ТЕХНИК|TEH|TECH|ЭЛЕКТРОНИК|SMART|DIGITAL|КОМПЬЮТЕР|IT|СВЯЗ|ТЕЛЕКОМ|MEGAFON",
  ("EL", "Электроника, техника и телеком")),
 (r"ОПТ|OPT|ДИСТРИБ|ТД |ТОРГОВЫЙ ДОМ|ТРЕЙД|TRADE|СНАБ|ГК |КОМПАНИЯ|ТК ",
  ("B2B", "Опт и мультикатегорийная дистрибуция")),
]
HEUR_C = [(re.compile(p, re.I), v) for p, v in HEUR]


def classify(domain, partner):
    if domain in KNOWN:
        return KNOWN[domain]
    blob = f"{partner} {domain}".upper()
    for rx, v in HEUR_C:
        if rx.search(blob):
            return v
    return ("?", "профиль не определён")


def load_partners():
    from openpyxl import load_workbook
    ws = load_workbook(f"{ROOT}/research/user_sites.xlsx", data_only=True)["Лист_1"]
    out = []
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, 6)]
        if not any(v):
            continue
        out.append(dict(
            partner=str(v[0] or "").strip(), contra=str(v[1] or "").strip(),
            inn=str(v[2]).strip() if v[2] else "",
            site=str(v[3]).strip() if v[3] else "", shop=str(v[4]).strip() if v[4] else ""))
    return out


def build_index():
    """domain -> запись мониторинга"""
    idx = {}

    # A. мой реестр
    for k in REG.S:
        r = REG.S[k]
        d = base(dom(r[1])) or base(dom(r[10]))
        if not d:
            continue
        idx[d] = dict(domain=d, name=r[0], code=r[2], sub=r[3], src="Мой поиск",
                      inn="", partners=[], brands=r[5], pl=r[6], region=r[7],
                      mech=r[8], url=r[10] if r[10].startswith("http") else f"https://{d}",
                      status=r[11], shop="")

    # B. партнёрский список
    for p in load_partners():
        ds = {base(dom(p["site"])), base(dom(p["shop"]))} - {""}
        has_shop = bool(dom(p["shop"]))
        for d in ds:
            if d in idx:
                e = idx[d]
                if e["src"] == "Мой поиск":
                    e["src"] = "Оба источника"
                e["inn"] = e["inn"] or p["inn"]
                e["partners"].append(p["partner"])
                if has_shop:
                    e["shop"] = "да"
            else:
                code, sub = classify(d, p["partner"])
                idx[d] = dict(domain=d, name=p["partner"] or d, code=code, sub=sub,
                              src="Партнёрский список", inn=p["inn"], partners=[p["partner"]],
                              brands="—", pl="—", region="Не проверялось",
                              mech="Домен из партнёрского списка; витрина не разбиралась",
                              url=f"https://{d}", status="Новый — требует разбора",
                              shop="да" if has_shop else "нет")
    return idx


# ---------------------------------------------------------------- сборка книги
GAP_ANALYSIS = [
 ("1. Смещение в «голову» рынка",
  "Я искал по запросам «купить батарейки», «элементы питания каталог». Поисковая выдача по таким "
  "запросам возвращает крупный ритейл и SEO-оптимизированные магазины.",
  "Найдены Ozon, Магнит, DNS, Лемана. Пропущены 340 доменов длинного хвоста.",
  "Идти от структуры рынка (реестр юрлиц, ОКВЭД, отраслевые каталоги), а не от поисковых запросов."),
 ("2. Название юрлица не равно домену",
  "Сопоставление «юрлицо → сайт» по названию не работает: у 340 компаний домен не выводится из названия.",
  "ООО «Альянс» → cellfaktor.ru; ГК «Движение» → gkm25.ru; ООО «Юста» → usta.group; "
  "ООО «ПАСКЕР ЛТД» → pasker.ru. Ровно поэтому 387 записей остались «не опознан».",
  "Резолвить домен по ИНН через ЕГРЮЛ/Контур.Фокус/DaData, а не по названию."),
 ("3. Батарейка — сопутствующая категория",
  "У большинства партнёров батарейки — одна из сотен позиций. Такие сайты не ранжируются "
  "по батарейным запросам вообще.",
  "elektro.ru, eltsnab.ru, informat.ru, megapack-russia.ru, retailhoreca.ru — "
  "электрика, канцелярия, упаковка, HoReCa.",
  "Искать не «кто продаёт батарейки», а «кто закупает у дистрибьюторов элементов питания»."),
 ("4. Неочевидные каналы вне ритейл-логики",
  "Целые сегменты не приходят в голову при поиске «где продаются батарейки».",
  "Аптеки (apteka.ru, zdravcity.ru, pharmeconom.ru), duty free (dme-dutyfree.ru), "
  "парфюмерия (goldapple.ru), охранные системы (bolid.ru, rubezh.ru), "
  "товары для взрослых (eroticfantasy.ru), ТВ-ритейл (leomax.ru).",
  "Строить карту каналов от потребления батарейки в устройстве, а не от полки магазина."),
 ("5. Региональный длинный хвост",
  "Поисковый индекс, доступный в этой сессии, помечен как US-only и слабо ранжирует "
  "региональные российские сайты.",
  "electrica-penza.ru, elektronika126.ru (Ставрополь), galaktika36.ru (Воронеж), "
  "narodny38.ru (Иркутск), krasyar.ru (Красноярск).",
  "Обходить региональные агрегаторы (2ГИС, Яндекс.Карты) и отраслевые B2B-каталоги по регионам."),
 ("6. Блокировка краулера",
  "Крупные сайты отдают 401/403/503 автоматическим запросам, что мешает подтверждать витрину.",
  "Лемана ПРО — 401, ЧипДип — 403, Комус — 503, Fix Price — 403, Авито и AliExpress закрыты полностью.",
  "Работать через headless-браузер (в этой среде предустановлен Chromium с Playwright), "
  "а не через простой HTTP-фетч."),
 ("7. Односторонняя проверка полноты",
  "У меня не было эталона, с которым сверить результат. Полнота не измерялась, "
  "и 103 площадки выглядели правдоподобно.",
  "Реальное покрытие оказалось 34 из 374 — 9%.",
  "Всегда фиксировать источник истины (партнёрский реестр, выгрузка 1С) и считать recall против него."),
]


def build(path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    idx = build_index()
    rows = sorted(idx.values(), key=lambda v: ([c[0] for c in CLASSIFIER].index(v["code"])
                                               if v["code"] in CLS_NAME else 99, v["domain"]))
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    body = Font(name=FONT, size=10)
    link = Font(name=FONT, size=10, color="0563C1", underline="single")
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    src_fill = {"Мой поиск": "E2EFDA", "Оба источника": "FFF2CC", "Партнёрский список": "FDE9E9"}

    def head(ws, hs, ws_, h=40):
        for c, t in enumerate(hs, 1):
            x = ws.cell(row=1, column=c, value=t)
            x.fill, x.font, x.border = hdr_fill, hdr_font, bd
            x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c, w in enumerate(ws_, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = h

    wb = Workbook()

    # --- Мониторинг ---
    ws = wb.active
    ws.title = "Мониторинг"
    head(ws, ["№", "Домен", "Партнёр / площадка", "ИНН", "Код", "Направление", "Подкатегория",
              "Источник", "Юрлиц", "Интернет-магазин", "Регион РФ / СНГ",
              "Выбор региона", "Ссылка", "Статус"],
         [5, 30, 40, 14, 7, 32, 34, 20, 8, 15, 14, 20, 34, 34])
    r = 2
    for i, v in enumerate(rows, 1):
        vals = [i, v["domain"], v["name"], v["inn"], v["code"],
                CLS_NAME.get(v["code"], "не определено"), v["sub"], v["src"],
                len(v["partners"]) or ("" if v["src"] == "Мой поиск" else 1),
                v["shop"] or "—", "СНГ" if is_cis(v["domain"]) else "РФ",
                v["region"], v["url"], v["status"]]
        f = PatternFill("solid", fgColor=src_fill[v["src"]])
        for c, val in enumerate(vals, 1):
            x = ws.cell(row=r, column=c, value=val)
            x.font, x.fill, x.border = body, f, bd
            x.alignment = Alignment(vertical="top", wrap_text=(c in (3, 6, 7, 12, 14)))
        lc = ws.cell(row=r, column=13); lc.hyperlink = v["url"]; lc.font = link
        lc.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    last = r - 1
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:N{last}"

    # --- Разрыв методологии ---
    ws2 = wb.create_sheet("Разрыв методологии")
    ws2.cell(row=1, column=1, value="Почему первый проход нашёл 34 из 374 доменов (9%)").font = \
        Font(name=FONT, size=12, bold=True)
    for c, t in enumerate(["Причина", "В чём была ошибка", "Примеры пропущенного", "Как исправить"], 1):
        x = ws2.cell(row=3, column=c, value=t)
        x.fill, x.font, x.border = hdr_fill, hdr_font, bd
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (a, b, c_, d) in enumerate(GAP_ANALYSIS, 4):
        for c, val in enumerate([a, b, c_, d], 1):
            x = ws2.cell(row=i, column=c, value=val)
            x.font, x.border = body, bd
            x.alignment = Alignment(vertical="top", wrap_text=True)
    for c, w in enumerate([34, 58, 58, 58], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[3].height = 30

    # --- Классификатор ---
    ws3 = wb.create_sheet("Классификатор")
    head(ws3, ["Код", "Направление", "Что относим", "Примеры"], [8, 40, 62, 40], 30)
    for i, (code, nm, desc, ex) in enumerate(CLASSIFIER, 2):
        for c, val in enumerate([code, nm, desc, ex], 1):
            x = ws3.cell(row=i, column=c, value=val)
            x.font, x.border = body, bd
            x.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Сводка ---
    ws4 = wb.create_sheet("Сводка")
    ws4.cell(row=1, column=1, value="Итог по объединённому реестру").font = Font(name=FONT, size=12, bold=True)
    for c, t in enumerate(["Направление", "Код", "Доменов"], 1):
        x = ws4.cell(row=3, column=c, value=t); x.fill, x.font, x.border = hdr_fill, hdr_font, bd
    rr = 4
    for code, nm, _, _ in CLASSIFIER:
        ws4.cell(row=rr, column=1, value=nm).font = body
        ws4.cell(row=rr, column=2, value=code).font = body
        ws4.cell(row=rr, column=3, value=f"=COUNTIF(Мониторинг!$E$2:$E${last},B{rr})").font = body
        for c in range(1, 4):
            ws4.cell(row=rr, column=c).border = bd
        rr += 1
    ws4.cell(row=rr, column=1, value="Профиль не определён").font = body
    ws4.cell(row=rr, column=2, value="?").font = body
    ws4.cell(row=rr, column=3, value=f'=COUNTIF(Мониторинг!$E$2:$E${last},"?")').font = body
    rr += 1
    ws4.cell(row=rr, column=1, value="ВСЕГО ДОМЕНОВ").font = Font(name=FONT, size=10, bold=True)
    ws4.cell(row=rr, column=3, value=f"=SUM(C4:C{rr-1})").font = Font(name=FONT, size=10, bold=True)

    rr += 2
    ws4.cell(row=rr, column=1, value="Источник домена").font = Font(name=FONT, size=11, bold=True)
    rr += 1
    for lbl in ["Мой поиск", "Оба источника", "Партнёрский список"]:
        ws4.cell(row=rr, column=1, value=lbl).font = body
        ws4.cell(row=rr, column=3, value=f"=COUNTIF(Мониторинг!$H$2:$H${last},A{rr})").font = body
        rr += 1
    rr += 1
    ws4.cell(row=rr, column=1, value="Полнота первого прохода (recall)").font = Font(name=FONT, size=11, bold=True)
    rr += 1
    for lbl, formula in [
        ("Доменов в партнёрском списке", f'=COUNTIF(Мониторинг!$H$2:$H${last},"Оба источника")'
                                        f'+COUNTIF(Мониторинг!$H$2:$H${last},"Партнёрский список")'),
        ("Из них нашёл первым проходом", f'=COUNTIF(Мониторинг!$H$2:$H${last},"Оба источника")'),
    ]:
        ws4.cell(row=rr, column=1, value=lbl).font = body
        ws4.cell(row=rr, column=3, value=formula).font = body
        rr += 1
    ws4.cell(row=rr, column=1, value="Recall").font = Font(name=FONT, size=10, bold=True)
    ws4.cell(row=rr, column=3, value=f"=IF(C{rr-2}=0,0,C{rr-1}/C{rr-2})").font = Font(name=FONT, size=10, bold=True)
    ws4.cell(row=rr, column=3).number_format = "0.0%"
    for c, w in enumerate([44, 10, 14], 1):
        ws4.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)
    return len(rows), last


if __name__ == "__main__":
    n, last = build(f"{ROOT}/RU_battery_monitoring.xlsx")
    print(f"сохранено: доменов {n}")
