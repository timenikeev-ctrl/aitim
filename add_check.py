# -*- coding: utf-8 -*-
"""Добавляет результаты живой проверки доменов в книгу мониторинга."""
import json, collections
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/home/user/aitim"
FONT = "Arial"
PATH = f"{ROOT}/RU_battery_monitoring.xlsx"

chk = {c["domain"]: c for c in json.load(open(f"{ROOT}/research/site_check.json"))}
wb = load_workbook(PATH)
ws = wb["Мониторинг"]

hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
body = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)

NEW = ["Сайт жив", "HTTP", "Батарейки на сайте", "Чем подтверждено", "Доказательство", "Title сайта"]
start = ws.max_column + 1
for i, t in enumerate(NEW):
    c = ws.cell(row=1, column=start + i, value=t)
    c.fill, c.font, c.border = hdr_fill, hdr_font, bd
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

FILL = {"да": "C6EFCE", "не найдено": "FFEB9C", "закрыт ботозащитой": "FFE0B2",
        "магазин, батареек не найдено": "FFEB9C", "не интернет-магазин": "E7E6E6",
        "недоступен через прокси": "DDEBF7", "сертификат не проходит проверку": "FCE4D6",
        "сайт не отвечает": "FFC7CE"}

for r in range(2, ws.max_row + 1):
    d = ws.cell(row=r, column=2).value
    c = chk.get(d)
    if not c:
        continue
    bat = c["battery"]
    vals = ["да" if c["alive"] else "нет", str(c["http"]), bat,
            c["method"] or "—", c["evidence"] or "—", c["title"] or "—"]
    fill = PatternFill("solid", fgColor=FILL.get(bat, "FFC7CE"))
    for i, v in enumerate(vals):
        x = ws.cell(row=r, column=start + i, value=v)
        x.font, x.border, x.fill = body, bd, fill
        x.alignment = Alignment(vertical="top", wrap_text=(i in (4, 5)))

for i, w in enumerate([10, 9, 22, 20, 52, 46]):
    ws.column_dimensions[get_column_letter(start + i)].width = w
ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

# --- лист с итогом проверки ---
if "Проверка сайтов" in wb.sheetnames:
    del wb["Проверка сайтов"]
wsc = wb.create_sheet("Проверка сайтов")
wsc.cell(row=1, column=1, value="Живая проверка всех доменов реестра").font = Font(name=FONT, size=12, bold=True)
wsc.cell(row=2, column=1,
         value="Каждый домен опрошен по HTTP: главная, robots.txt/sitemap, поиск по сайту, "
               "типовые URL категорий. Chromium через прокси этой среды не проходит, "
               "поэтому сайты с ботозащитой отмечены отдельно, а не записаны в мёртвые.").font = \
    Font(name=FONT, size=10, italic=True)

rows = [("Батарейки подтверждены", "да"),
        ("Магазин есть, батареек не найдено", "магазин, батареек не найдено"),
        ("Не интернет-магазин (сайт-визитка)", "не интернет-магазин"),
        ("Недоступен через прокси среды", "недоступен через прокси"),
        ("Сертификат не проходит проверку", "сертификат не проходит проверку"),
        ("Закрыт ботозащитой (подтверждено)", "закрыт ботозащитой")]
for i, t in enumerate(["Результат", "Доменов", "Доля"], 1):
    x = wsc.cell(row=4, column=i, value=t)
    x.fill, x.font, x.border = hdr_fill, hdr_font, bd
cnt = collections.Counter(c["battery"] for c in chk.values())
dead = sum(v for k, v in cnt.items() if k.startswith("сайт не отвечает") or k.startswith("ошибка"))
total = len(chk)
data = [(rows[0][0], cnt.get("да", 0)),
        (rows[1][0], cnt.get("магазин, батареек не найдено", 0)),
        (rows[2][0], cnt.get("не интернет-магазин", 0)),
        (rows[3][0], cnt.get("недоступен через прокси", 0)),
        (rows[4][0], cnt.get("сертификат не проходит проверку", 0)),
        (rows[5][0], cnt.get("закрыт ботозащитой", 0))]
r = 5
for label, n in data:
    wsc.cell(row=r, column=1, value=label).font = body
    wsc.cell(row=r, column=2, value=n).font = body
    wsc.cell(row=r, column=3, value=f"=B{r}/$B${r + len(data) - (r - 5)}").font = body
    for c in range(1, 4):
        wsc.cell(row=r, column=c).border = bd
    r += 1
wsc.cell(row=r, column=1, value="ВСЕГО").font = Font(name=FONT, size=10, bold=True)
wsc.cell(row=r, column=2, value=f"=SUM(B5:B{r-1})").font = Font(name=FONT, size=10, bold=True)
for rr in range(5, r):
    wsc.cell(row=rr, column=3).value = f"=B{rr}/$B${r}"
    wsc.cell(row=rr, column=3).number_format = "0.0%"
for c, w in enumerate([40, 12, 10], 1):
    wsc.column_dimensions[get_column_letter(c)].width = w

# метод подтверждения
r += 2
wsc.cell(row=r, column=1, value="Чем подтверждено наличие батареек").font = Font(name=FONT, size=11, bold=True)
r += 1
for i, t in enumerate(["Метод", "Доменов"], 1):
    x = wsc.cell(row=r, column=i, value=t); x.fill, x.font, x.border = hdr_fill, hdr_font, bd
r += 1
for k, v in collections.Counter(c["method"] for c in chk.values() if c["battery"] == "да").most_common():
    wsc.cell(row=r, column=1, value=k).font = body
    wsc.cell(row=r, column=2, value=v).font = body
    wsc.cell(row=r, column=1).border = bd; wsc.cell(row=r, column=2).border = bd
    r += 1

wb.save(PATH)
print(f"обновлено. Проверено {total}: подтверждено {cnt.get('да',0)}, "
      f"не найдено {cnt.get('не найдено',0)}, ботозащита {cnt.get('закрыт ботозащитой',0)}, мёртвых {dead}")
