# -*- coding: utf-8 -*-
"""Сборка реестра российских онлайн-площадок, продающих батарейки (элементы питания)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = "Arial"

# --- Классификатор направлений деятельности (разработан для этой задачи) ---
CLASSIFIER = [
    ("MP",   "Маркетплейс",
     "Универсальная площадка с внешними продавцами (3P), категория ведётся селлерами",
     "Ozon, Wildberries, Яндекс Маркет"),
    ("EL",   "Электроника и бытовая техника",
     "Специализированный ритейл техники; батарейки — сопутствующая категория/аксессуар",
     "DNS, Ситилинк, М.Видео"),
    ("FMCG", "Продукты и FMCG-ритейл",
     "Продуктовые сети (гипер-/супермаркеты), батарейки в прикассовой и нон-фуд зоне",
     "Магнит, Лента, METRO"),
    ("DIY",  "DIY, ремонт и товары для дома",
     "Стройматериалы, электрика, инструмент; батарейки в электротоварах",
     "Лемана ПРО, ВсеИнструменты, Петрович"),
    ("DROG", "Дрогери и фикс-прайс",
     "Нон-фуд FMCG с низким средним чеком, высокая доля СТМ",
     "Fix Price, Улыбка радуги, Галамарт"),
    ("B2B",  "Офис, канцелярия и B2B-опт",
     "Снабжение юрлиц, оптовые каталоги, прайс-листы, отсрочка платежа",
     "Комус, Самсон-опт, ЧипДип"),
    ("SPEC", "Специализированные магазины элементов питания",
     "Моно-категорийные игроки: только батарейки, аккумуляторы, ЗУ",
     "Мир Батареек, Склад батареек, Баттерика"),
    ("QCOM", "Q-commerce и e-grocery доставка",
     "Быстрая доставка продуктов и товаров первой необходимости, витрина привязана к дарксторy/магазину",
     "Купер, Впрок, Самокат"),
]

CLS_NAME = {c[0]: c[1] for c in CLASSIFIER}

# Колонки:
# продавец, домен, код, тип игрока, бренды, СТМ, регион(ответ), механика, пример рег. URL, ссылка, статус
ROWS = [
    # ---------------- MP ----------------
    ("Ozon", "ozon.ru", "MP", "Маркетплейс 1P+3P",
     "Весь спектр: GP, Duracell, Energizer, Varta, Panasonic, Camelion, Kodak, Robiton, SmartBuy, ЭРА",
     "Нет (есть Ozon Basic у части селлеров)",
     "Единая карточка",
     "Один URL карточки на всю РФ. Регион меняет цену, срок и наличие, но не адрес страницы",
     "—",
     "https://www.ozon.ru/category/batareyki-15882/",
     "Проверено"),
    ("Wildberries", "wildberries.ru", "MP", "Маркетплейс 3P",
     "Весь спектр: Duracell, Energizer, GP, Varta, Camelion, SmartBuy, Perfeo, ЭРА",
     "Нет",
     "Единая карточка",
     "Единый артикул на РФ; регион влияет на срок доставки до ПВЗ, URL неизменен",
     "—",
     "https://www.wildberries.ru/catalog/elektronika/tags/elementy-pitanija",
     "Проверено"),
    ("Яндекс Маркет", "market.yandex.ru", "MP", "Маркетплейс 3P + витрина сравнения",
     "Duracell, Energizer, GP, Kodak, Commo, Sonnen, Perfeo и др.",
     "Commo (СТМ Яндекса)",
     "Единая карточка",
     "Карточка модели общая на РФ, региональность — в блоке офферов и сроках доставки",
     "—",
     "https://market.yandex.ru/catalog--batareiki-i-akkumuliatory-dlia-audio-i-videotekhniki/26914470/list",
     "Проверено"),
    ("Мегамаркет", "megamarket.ru", "MP", "Маркетплейс 3P (экосистема Сбер)",
     "GP, Duracell, Energizer, Varta, АШАН Красная птица, Magnit и др.",
     "Нет",
     "Единая карточка",
     "Единый URL категории и карточки, регион переключается селектором",
     "—",
     "https://megamarket.ru/catalog/batarejki/",
     "Проверено"),
    ("AliExpress Россия", "aliexpress.ru", "MP", "Маркетплейс 3P (локальные + трансграничные)",
     "Преимущественно азиатские марки + GP, Panasonic, SmartBuy",
     "Нет",
     "Единая карточка",
     "Единый URL товара, регион влияет на доступность локального склада",
     "—",
     "https://aliexpress.ru/",
     "Требует уточнения (домен не открылся для проверки)"),
    ("Авито", "avito.ru", "MP", "Классифайд + маркетплейс товаров",
     "Разрозненный ассортимент от продавцов",
     "Нет",
     "Есть версии",
     "Объявление публикуется в конкретном городе; выдача и URL включают гео-сегмент",
     "avito.ru/moskva/... vs avito.ru/sankt-peterburg/...",
     "https://www.avito.ru/",
     "Не проверено (сайт закрыт для краулера)"),

    # ---------------- EL ----------------
    ("DNS", "dns-shop.ru", "EL", "Федеральная сеть техники (омниканал)",
     "Duracell, GP, Energizer, Varta, Panasonic, Camelion, Robiton, SmartBuy, Kodak",
     "Есть (собственные марки DNS)",
     "Единая карточка",
     "Один URL на РФ, город выбирается селектором; цена и наличие отличаются по городам",
     "—",
     "https://www.dns-shop.ru/catalog/17a88ded16404e77/batarejki/",
     "Проверено"),
    ("Ситилинк", "citilink.ru", "EL", "Федеральная сеть техники (онлайн + ПВЗ)",
     "Duracell, VARTA, GP, Energizer, Camelion, SmartBuy",
     "Есть",
     "Единая карточка",
     "Единый URL категории/бренда, город в заголовке страницы (напр. «— Москва»)",
     "—",
     "https://www.citilink.ru/catalog/batareiki/",
     "Проверено"),
    ("М.Видео", "mvideo.ru", "EL", "Федеральная сеть техники (омниканал)",
     "GP, Duracell, Energizer, Varta, Panasonic, Kodak, Carrera",
     "Carrera (СТМ группы М.Видео-Эльдорадо)",
     "Единая карточка",
     "Единый URL, фильтр по бренду в query; регион — селектором",
     "—",
     "https://www.mvideo.ru/aksessuary-dlya-doma-50/batareiki-i-akkumulyatory-296",
     "Проверено"),
    ("Эльдорадо", "eldorado.ru", "EL", "Федеральная сеть техники (омниканал)",
     "GP Extra Alkaline, Kodak Xtralife, Energizer, Carrera, CROMEX",
     "Carrera (СТМ группы М.Видео-Эльдорадо)",
     "Единая карточка",
     "Единый URL категории, регион переключается селектором",
     "—",
     "https://www.eldorado.ru/c/batareyki/",
     "Проверено"),
    ("Технопарк", "technopark.ru", "EL", "Сеть техники (Москва + онлайн по РФ)",
     "GP, Duracell, Energizer, Varta, Camelion",
     "Нет",
     "Единая карточка",
     "Единый URL раздела «Элементы питания»",
     "—",
     "https://www.technopark.ru/elementy-pitaniya/",
     "Проверено"),
    ("ОНЛАЙН ТРЕЙД.РУ", "onlinetrade.ru", "EL", "Онлайн-ритейлер техники + ПВЗ",
     "VARTA, GP, Duracell, Panasonic, Energizer, Robiton",
     "Нет",
     "Единая карточка",
     "Единый URL карточки, регион — селектором города/ПВЗ",
     "—",
     "https://www.onlinetrade.ru/catalogue/batareyki-c59/",
     "Проверено"),
    ("Регард", "regard.ru", "EL", "Онлайн-ритейлер техники (ядро — Москва)",
     "GP, Duracell, Varta, Energizer, Camelion",
     "Нет",
     "Единая карточка",
     "Единый каталог, привязка к Москве",
     "—",
     "https://www.regard.ru/catalog/1223/batareiki-akkumuliatory",
     "Проверено"),
    ("RBT.ru", "rbt.ru", "EL", "Сеть техники (Урал, Сибирь, юг РФ)",
     "GP, Duracell, Energizer, Varta, Camelion",
     "Нет",
     "Есть версии",
     "Полноценные региональные поддомены: у каждого города свой сайт и свой URL категории",
     "www.rbt.ru (Челябинск) vs blagodarniy.rbt.ru",
     "https://www.rbt.ru/cat/tele-video-audio/batareyki/",
     "Проверено"),

    # ---------------- FMCG ----------------
    ("Магнит", "magnit.ru", "FMCG", "Крупнейшая продуктовая сеть РФ (омниканал)",
     "GP, Duracell, Energizer, Varta + СТМ",
     "Есть (СТМ «Магнит»)",
     "Есть версии",
     "Каталог привязан к конкретной торговой точке: ID магазина зашит в URL категории",
     "magnit.ru/catalog/53067-batareyki vs magnit.ru/catalog/107358-batareyki_1 (Краснодар, Дзержинского 42)",
     "https://www.magnit.ru/catalog/53067-batareyki",
     "Проверено"),
    ("Магнит Косметик", "cosmetic.magnit.ru", "DROG", "Дрогери-формат группы «Магнит»",
     "GP Super, Duracell, Energizer + СТМ",
     "Есть (СТМ «Магнит»)",
     "Единая карточка",
     "Отдельный поддомен баннера; URL товара стабилен, город подставляется в контексте",
     "cosmetic.magnit.ru (город в шапке, напр. Краснодар)",
     "https://cosmetic.magnit.ru/product/4000085367-batareyki-gp-super-aa-6sht",
     "Проверено"),
    ("Пятёрочка", "5ka.ru", "FMCG", "Дискаунтер у дома (X5 Group)",
     "GP, Duracell, Energizer, Varta + СТМ",
     "Есть (СТМ X5)",
     "Единая карточка",
     "Единый URL категории; витрина и цена подтягиваются под выбранный магазин доставки",
     "—",
     "https://5ka.ru/catalog/batareyki--251C12959/",
     "Проверено"),
    ("Перекрёсток", "perekrestok.ru", "FMCG", "Супермаркеты (X5 Group)",
     "Duracell, GP, Energizer, Varta + СТМ",
     "Есть (СТМ X5)",
     "Единая карточка",
     "Единый URL, город и магазин — селектором",
     "—",
     "https://www.perekrestok.ru/cat/search?search=%D0%B1%D0%B0%D1%82%D0%B0%D1%80%D0%B5%D0%B9%D0%BA%D0%B8",
     "Проверено"),
    ("Лента", "lenta.com", "FMCG", "Гипер- и супермаркеты (Лента Групп)",
     "Duracell, Varta, Energizer, GP, «365 дней»",
     "Есть («365 дней»)",
     "Единая карточка",
     "Единый URL категории на РФ, город выбирается селектором",
     "—",
     "https://lenta.com/catalog/vse-dlya-doma/lampochki-elementy-pitaniya/elementy-pitaniya/",
     "Проверено"),
    ("METRO Cash & Carry", "online.metro-cc.ru", "FMCG", "Мелкооптовый гипермаркет (B2B+B2C)",
     "Duracell, Energizer, GP, Varta, Panasonic + СТМ",
     "Есть (Aro, Rioba и др.)",
     "Единая карточка",
     "Единый URL, торговый центр и город — селектором",
     "—",
     "https://online.metro-cc.ru/category/tovary-dlya-doma-dachi-sada/elektronika-i-tekhnika/batareyki-zaryadnye-ustroystva-fonariki",
     "Проверено"),
    ("Глобус", "online.globus.ru", "FMCG", "Гипермаркеты (Москва и ЦФО)",
     "Varta CR2032, Energizer, GP, Duracell",
     "Есть",
     "Единая карточка",
     "Единый URL категории, витрина под выбранный гипермаркет",
     "—",
     "https://online.globus.ru/catalog/dom-khobbi-tekhnika/elektrotovary/batareyki-i-akkumulyatory/",
     "Проверено"),
    ("АШАН", "auchan.ru", "FMCG", "Гипермаркеты + онлайн-доставка",
     "GP, Duracell, Energizer, Varta, АШАН Красная птица",
     "Есть («АШАН Красная птица»)",
     "Единая карточка",
     "Единый URL категории; доставка в МСК, СПб, Ростов, Самара, Н.Новгород, Новосибирск, Краснодар, Татарстан и др.",
     "—",
     "https://www.auchan.ru/catalog/stroitelstvo-i-remont/elementy-pitaniya-i-fonari/batareyki/",
     "Проверено"),
    ("О'КЕЙ", "okeydostavka.ru", "FMCG", "Гипермаркеты и супермаркеты (25 городов)",
     "Duracell, GP, Energizer, Varta + СТМ",
     "Есть (СТМ «О'КЕЙ»)",
     "Есть версии",
     "Регион зашит в путь URL — отдельная витрина на город",
     "okeydostavka.ru/msk/... (Москва); аналогичные сегменты для других городов",
     "https://www.okeydostavka.ru/msk/tovary-dlia-doma/elektrotovary/batareiki",
     "Проверено"),

    # ---------------- DIY ----------------
    ("Лемана ПРО (быв. Леруа Мерлен)", "lemanapro.ru", "DIY", "DIY-гипермаркеты, федеральная сеть",
     "Duracell, Camelion, GP, Varta, Energizer, Космос",
     "Есть (Lemana / СТМ сети)",
     "Есть версии",
     "Классические региональные поддомены: у каждого города свой сайт с собственным URL категории",
     "lemanapro.ru (Москва) / spb.lemanapro.ru / irkutsk.lemanapro.ru",
     "https://lemanapro.ru/catalogue/batareyki/",
     "Проверено"),
    ("ВсеИнструменты.ру", "vseinstrumenti.ru", "DIY", "Онлайн-DIY №1, 306 магазинов/ПВЗ",
     "Duracell, GP, Varta, Energizer, Panasonic, КОСМОС, ТРОФИ, ФАZА, Robiton, SmartBuy, Camelion, GoPower, Navigator",
     "Есть (ТРОФИ, ФАZА и др.)",
     "Единая карточка",
     "Единый домен, город выбирается пикером; ~4250 SKU в категории «Батарейки»",
     "—",
     "https://www.vseinstrumenti.ru/category/batarejki-4645/",
     "Проверено"),
    ("Петрович", "petrovich.ru", "DIY", "Строительные центры (СПб, Москва, СЗФО, ЦФО)",
     "GP, Philips, VARTA, Duracell",
     "Есть",
     "Есть версии",
     "Региональные поддомены с раздельными карточками и ценами",
     "petrovich.ru (СПб) vs moscow.petrovich.ru",
     "https://petrovich.ru/catalog/285395742/",
     "Проверено"),
    ("Строительный двор", "sdvor.com", "DIY", "Стройматериалы, 14 городов (Урал, Тюмень, Москва)",
     "GP, Duracell, Varta, Космос",
     "Есть",
     "Есть версии",
     "Код города в пути URL — отдельная витрина и цена на каждый город",
     "sdvor.com/tmn/... (Тюмень) vs sdvor.com/ekb/... (Екатеринбург)",
     "https://www.sdvor.com/tmn/category/batarejki-5826",
     "Проверено"),
    ("Максидом", "maxidom.ru", "DIY", "DIY-гипермаркеты (СПб, Москва, регионы)",
     "GP, Duracell, Varta, Energizer, Camelion",
     "Есть",
     "Единая карточка",
     "Единый URL категории, город — селектором",
     "—",
     "https://www.maxidom.ru/catalog/batareyki/",
     "Проверено"),
    ("220 Вольт", "220-volt.ru", "DIY", "Инструмент и электротовары, онлайн + сеть",
     "GP, Duracell, КОСМОС, TOSHIBA, ЗУБР, Fenix",
     "Нет",
     "Единая карточка",
     "Единый URL, доставка «Москва, СПб и РФ»; есть отдельный оптовый раздел",
     "—",
     "https://www.220-volt.ru/catalog/akkumulyatory-batareiki/",
     "Проверено"),
    ("Дом Лента (быв. OBI Россия)", "obi.ru", "DIY", "DIY-гипермаркеты, 23 объекта в 11 городах",
     "GP, Duracell, Varta, Energizer",
     "Есть (СТМ Ленты)",
     "Требует уточнения",
     "Сеть вошла в «Ленту», с 2026 идёт поэтапный ребрендинг в «Дом Лента» — структура сайта меняется",
     "—",
     "https://www.obi.ru/",
     "Статус меняется (ребрендинг 2026)"),

    # ---------------- DROG ----------------
    ("Fix Price", "fix-price.com", "DROG", "Сеть магазинов фиксированных цен",
     "FLARX (СТМ), безымянные солевые/алкалиновые",
     "Есть (FLARX)",
     "Единая карточка",
     "Единый URL товара, город подставляется в контексте («в г. Москва»)",
     "—",
     "https://fix-price.com/catalog/dlya-doma/p-5014027-batareyki-alkalinovye-aa-4-sht",
     "Проверено"),
    ("Улыбка радуги", "r-ulybka.ru", "DROG", "Дрогери, 1000+ магазинов",
     "GP, Duracell, Energizer, Camelion + СТМ",
     "Есть",
     "Единая карточка",
     "Единый URL категории, доставка в Москву, СПб и другие города",
     "—",
     "https://www.r-ulybka.ru/catalog/batarejki-i-lampochki/",
     "Проверено"),
    ("Галамарт", "galamart.ru", "DROG", "Сеть постоянных распродаж (нон-фуд)",
     "Camelion, GP, СТМ",
     "Есть",
     "Единая карточка",
     "Единый URL категории «Батарейки»",
     "—",
     "https://galamart.ru/catalog/elektronika/batareiki/",
     "Проверено"),

    # ---------------- B2B ----------------
    ("Комус", "komus.ru", "B2B", "Снабжение офиса и дома, 4500+ пунктов выдачи",
     "GP, Duracell, Energizer, Varta, Комус (СТМ)",
     "Есть (ТМ «Комус»)",
     "Единая карточка",
     "Один домен; ассортимент, цена и условия доставки зависят от выбранного региона, URL не меняется",
     "—",
     "https://www.komus.ru/katalog/katalog-instrumentov/elektrika-i-svet/batarejki-akkumulyatory-zaryadnye-ustrojstva/batarejki/c/987489/",
     "Проверено"),
    ("ЧипДип", "chipdip.ru", "B2B", "Радиокомпоненты и электроника, опт + розница",
     "Duracell, Energizer, GP, Varta, Panasonic, Renata, Robiton, Camelion, промышленные Saft/Tadiran",
     "Есть (Rexant и партнёрские ТМ)",
     "Единая карточка",
     "Единый каталог, город — селектором",
     "—",
     "https://www.chipdip.ru/catalog/batteries",
     "Проверено"),
    ("Самсон-опт", "samsonopt.ru", "B2B", "Крупнейший оптовый поставщик товаров для офиса",
     "GP, DURACELL, SONNEN, CROMEX, BRAUBERG",
     "Есть (SONNEN, CROMEX, BRAUBERG)",
     "Единая карточка",
     "Оптовый каталог с единым URL; доступ по договору, отгрузка со склада по РФ",
     "—",
     "https://www.samsonopt.ru/zakaz/index.php?ID=25792",
     "Проверено"),
    ("ОФИСМАГ", "officemag.ru", "B2B", "Канцтовары и товары для офиса (розница + опт)",
     "GP, Duracell, SONNEN, CROMEX",
     "Есть (SONNEN, CROMEX)",
     "Единая карточка",
     "Единый URL категории с фильтрами по форме и типу батарейки",
     "—",
     "https://www.officemag.ru/catalog/1255/",
     "Проверено"),
    ("Сима-ленд", "sima-land.ru", "B2B", "Оптово-розничный универсальный маркет (Екатеринбург)",
     "GP, Duracell, Luazon, Camelion, ЭРА",
     "Есть (Luazon)",
     "Единая карточка",
     "Единый URL категории, цены опт/розница на одной странице",
     "—",
     "https://www.sima-land.ru/batareyki/",
     "Проверено"),

    # ---------------- SPEC ----------------
    ("Мир Батареек", "mirbatareek.ru", "SPEC", "Моно-категорийный магазин, опт + розница, с 2011 г.",
     "Duracell, Panasonic, Energizer, VARTA, GARIN, GP",
     "Нет",
     "Единая карточка",
     "Регионального селектора нет: один склад в Москве (ул. Красная сосна, 2), доставка по РФ",
     "—",
     "https://mirbatareek.ru/catalog/",
     "Проверено (прямой разбор страницы)"),
    ("Склад батареек", "zbat.ru", "SPEC", "Оптовый склад элементов питания (мин. заказ 500 ₽)",
     "Duracell, Energizer, Panasonic, Sony, Varta, Camelion, GP, Kodak, Robiton, Ansmann, TADIRAN, Saft, Renata, Maxell, Космос, Облик",
     "Нет",
     "Единая карточка",
     "Единый каталог; два хаба (Москва, СПб) и раздел «Регионы», но URL товара общий",
     "—",
     "https://zbat.ru/batarejki/",
     "Проверено (прямой разбор страницы)"),
    ("Баттерика", "batterika.ru", "SPEC", "Источники питания и комплектующие, розница",
     "ANSMANN, CAMELION, Duracell, ENERGIZER, GP, GoPower, Kodak, MAXELL, MIREX, Minamoto, Panasonic, Renata, Robiton, SMARTBUY, Saft, Toshiba, Varta, Vinnic, ЭРА",
     "Нет",
     "Единая карточка",
     "Есть селектор города (Москва, СПб, Новосибирск, Екатеринбург, Казань и др.), URL товара не меняется",
     "—",
     "https://batterika.ru/batareyki/",
     "Проверено (прямой разбор страницы)"),
    ("BatteryMag", "batterymag.ru", "SPEC", "Магазин батареек и аккумуляторов, опт + розница",
     "Duracell, Energizer, GP, Varta, Panasonic, Robiton",
     "Нет",
     "Единая карточка",
     "Единый каталог без региональных версий",
     "—",
     "https://batterymag.ru/",
     "Проверено"),
    ("Батарейки оптом (battery-opt.ru)", "battery-opt.ru", "SPEC", "Оптовый поставщик, мелкий и крупный опт по РФ",
     "Energizer, Panasonic, Varta, Duracell, GP",
     "Нет",
     "Единая карточка",
     "Единый каталог, разделение на бытовой и промышленный ассортимент",
     "—",
     "https://battery-opt.ru/pc/power-elements/",
     "Проверено"),
    ("Киликс (Qilix)", "qilix.ru", "SPEC", "Дистрибьютор элементов питания, опт",
     "Varta, SmartBuy, Saft, GP, Duracell, Energizer",
     "Нет",
     "Единая карточка",
     "Единый каталог с доставкой в любую точку РФ",
     "—",
     "https://qilix.ru/catalog/batareiki-optom/aa/",
     "Проверено"),
    ("ShoWay", "showay.ru", "SPEC", "Магазин батареек, опт + розница",
     "GP, Duracell, Energizer, Varta, Camelion",
     "Нет",
     "Единая карточка",
     "Единый каталог",
     "—",
     "https://showay.ru/cat/batareyki/",
     "Проверено"),
    ("Батарейки и элементы питания оптом", "batareiki-optom.ru", "SPEC", "Оптовый магазин, 23+ бренда, 1000+ SKU",
     "Tadiran, GoPower, Robiton, Varta, Duracell, Energizer, Panasonic",
     "Нет",
     "Единая карточка",
     "Единый каталог, самовывоз и доставка",
     "—",
     "https://batareiki-optom.ru/",
     "Проверено"),
    ("Все Батарейки", "всебатарейки.рф", "SPEC", "Интернет-магазин элементов питания и электротоваров",
     "Duracell, Energizer, GP, Varta, Camelion",
     "Нет",
     "Единая карточка",
     "Единый каталог",
     "—",
     "https://xn--80aabdpdvhj1dgp.xn--p1ai/",
     "Проверено"),
    ("Сеть магазинов «Батарейка»", "батарейка-сеть.рф", "SPEC", "Сеть специализированных офлайн-магазинов + сайт",
     "Duracell, Energizer, GP, Varta, Robiton",
     "Нет",
     "Требует уточнения",
     "Сайт сети; региональная логика витрины не подтверждена",
     "—",
     "https://xn----7sbabaaenxplrnqu6c6a.xn--p1ai/",
     "Требует уточнения"),

    # ---------------- QCOM ----------------
    ("Впрок (Перекрёсток Впрок)", "vprok.ru", "QCOM", "Онлайн-гипермаркет X5 (заказ на неделю)",
     "Duracell, GP, Energizer, Varta + СТМ X5",
     "Есть (СТМ X5)",
     "Единая карточка",
     "Единый URL категории; зона доставки ограничена регионами присутствия",
     "—",
     "https://www.vprok.ru/catalog/1458/batareyki",
     "Проверено"),
    ("Купер (быв. СберМаркет)", "kuper.ru", "QCOM", "Агрегатор доставки из сетей (Ашан, Лента, METRO, Магнит, Пятёрочка)",
     "Витрина наследует бренды сети-партнёра",
     "Наследуется от сети",
     "Есть версии",
     "Витрина строится под конкретный магазин-партнёр и город; у каждой сети свой раздел",
     "kuper.ru/auchan и аналогичные разделы партнёров",
     "https://kuper.ru/auchan",
     "Проверено"),
    ("Самокат", "samokat.ru", "QCOM", "Экспресс-доставка из даркстора",
     "СТМ и базовые марки (GP, Duracell)",
     "Есть (СТМ «Самокат»)",
     "Есть версии",
     "Даркстор-модель: витрина и наличие привязаны к ближайшему дарксторy; ~1100 дарксторов в 49 городах, ассортимент между ними различается",
     "—",
     "https://samokat.ru/",
     "Проверено"),

    # ---------------- Прочее ----------------
    ("Детский мир", "detmir.ru", "FMCG", "Сеть детских товаров (батарейки к игрушкам)",
     "GP, Duracell, Energizer, Varta, Camelion",
     "Есть",
     "Единая карточка",
     "Единый URL категории, доставка по РФ (Москва, СПб, Н.Новгород, Челябинск, Новосибирск, Тюмень и др.)",
     "—",
     "https://www.detmir.ru/catalog/index/name/batteries1/",
     "Проверено"),
]


def build(path):
    wb = Workbook()

    # ---------- Лист 1: Продавцы ----------
    ws = wb.active
    ws.title = "Продавцы"

    headers = [
        "№",
        "Продавец / торговая сеть",
        "Домен",
        "Код направления",
        "Направление деятельности",
        "Тип игрока",
        "Бренды элементов питания в ассортименте",
        "СТМ / собственная марка",
        "Выбор региона",
        "Как устроена региональность",
        "Пример региональных URL",
        "Ссылка на страницу «Батарейки»",
        "Статус проверки",
    ]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    body_font = Font(name=FONT, size=10)
    link_font = Font(name=FONT, size=10, color="0563C1", underline="single")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Заливка по направлению деятельности
    cat_fill = {
        "MP":   "E8F1FB",
        "EL":   "EAF3E8",
        "FMCG": "FFF6E5",
        "DIY":  "F2ECF7",
        "DROG": "FDECEC",
        "B2B":  "EAF6F6",
        "SPEC": "FFF9DB",
        "QCOM": "F0F0F0",
    }
    region_fill_versions = PatternFill("solid", fgColor="FFE699")
    region_font_versions = Font(name=FONT, size=10, bold=True, color="9C5700")

    r = 2
    for i, row in enumerate(ROWS, start=1):
        (name, domain, code, ptype, brands, pl, region, mech, rex, url, status) = row
        values = [i, name, domain, code, CLS_NAME[code], ptype, brands, pl, region, mech, rex, url, status]
        fill = PatternFill("solid", fgColor=cat_fill[code])
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (5, 6, 7, 8, 10, 11, 13)))
        # гиперссылка
        lc = ws.cell(row=r, column=12)
        lc.hyperlink = url
        lc.font = link_font
        lc.alignment = Alignment(vertical="top", wrap_text=True)
        # подсветка региональных версий
        if region == "Есть версии":
            rc = ws.cell(row=r, column=9)
            rc.fill = region_fill_versions
            rc.font = region_font_versions
        r += 1

    last = r - 1
    widths = [5, 30, 22, 9, 30, 34, 46, 26, 18, 46, 40, 52, 24]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:M{last}"

    # ---------- Лист 2: Классификатор ----------
    ws2 = wb.create_sheet("Классификатор")
    h2 = ["Код", "Направление деятельности", "Что относим", "Примеры"]
    for c, h in enumerate(h2, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for i, (code, nm, desc, ex) in enumerate(CLASSIFIER, start=2):
        for c, v in enumerate([code, nm, desc, ex], start=1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = body_font
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=cat_fill[code])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for c, w in enumerate([8, 40, 62, 40], start=1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 30

    note_row = len(CLASSIFIER) + 3
    ws2.cell(row=note_row, column=1,
             value="Классификация разработана под эту задачу: разделение по роли батарейки в ассортименте "
                   "(основная категория / сопутствующая / импульсная покупка) и по модели продаж (1P, 3P, опт).").font = \
        Font(name=FONT, size=10, italic=True)
    ws2.cell(row=note_row + 1, column=1,
             value="«Выбор региона»: «Есть версии» = у региона отдельный URL (поддомен, префикс пути или ID магазина). "
                   "«Единая карточка» = один URL на РФ, регион меняет только цену/наличие/сроки.").font = \
        Font(name=FONT, size=10, italic=True)

    # ---------- Лист 3: Сводка ----------
    ws3 = wb.create_sheet("Сводка")
    ws3.cell(row=1, column=1, value="Сводка по реестру").font = Font(name=FONT, size=12, bold=True)

    ws3.cell(row=3, column=1, value="Направление").font = hdr_font
    ws3.cell(row=3, column=2, value="Код").font = hdr_font
    ws3.cell(row=3, column=3, value="Площадок").font = hdr_font
    ws3.cell(row=3, column=4, value="Из них с региональными версиями").font = hdr_font
    for c in range(1, 5):
        cell = ws3.cell(row=3, column=c)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row3 = 4
    for code, nm, _, _ in CLASSIFIER:
        ws3.cell(row=row3, column=1, value=nm).font = body_font
        ws3.cell(row=row3, column=2, value=code).font = body_font
        ws3.cell(row=row3, column=3,
                 value=f'=COUNTIF(Продавцы!$D$2:$D${last},B{row3})').font = body_font
        ws3.cell(row=row3, column=4,
                 value=f'=COUNTIFS(Продавцы!$D$2:$D${last},B{row3},Продавцы!$I$2:$I${last},"Есть версии")').font = body_font
        for c in range(1, 5):
            ws3.cell(row=row3, column=c).border = border
            ws3.cell(row=row3, column=c).fill = PatternFill("solid", fgColor=cat_fill[code])
        row3 += 1

    ws3.cell(row=row3, column=1, value="ИТОГО").font = Font(name=FONT, size=10, bold=True)
    ws3.cell(row=row3, column=3, value=f"=SUM(C4:C{row3-1})").font = Font(name=FONT, size=10, bold=True)
    ws3.cell(row=row3, column=4, value=f"=SUM(D4:D{row3-1})").font = Font(name=FONT, size=10, bold=True)
    for c in range(1, 5):
        ws3.cell(row=row3, column=c).border = border
        ws3.cell(row=row3, column=c).fill = PatternFill("solid", fgColor="D9D9D9")

    # блок по региональности
    rr = row3 + 2
    ws3.cell(row=rr, column=1, value="Архитектура региональности").font = Font(name=FONT, size=11, bold=True)
    rr += 1
    for c, h in enumerate(["Ответ", "Площадок"], start=1):
        cell = ws3.cell(row=rr, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
    rr += 1
    for label in ["Единая карточка", "Есть версии", "Требует уточнения"]:
        ws3.cell(row=rr, column=1, value=label).font = body_font
        ws3.cell(row=rr, column=2,
                 value=f'=COUNTIF(Продавцы!$I$2:$I${last},A{rr})').font = body_font
        ws3.cell(row=rr, column=1).border = border
        ws3.cell(row=rr, column=2).border = border
        rr += 1

    for c, w in enumerate([44, 10, 14, 32], start=1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    rr += 1
    ws3.cell(row=rr, column=1,
             value="Источник данных: открытые каталоги площадок, сентябрь 2026. "
                   "Поле «Статус проверки» на листе «Продавцы» показывает, что подтверждено напрямую, "
                   "а что требует ручной перепроверки.").font = Font(name=FONT, size=9, italic=True)

    wb.save(path)
    print(f"saved: {path}, rows: {last-1}")


if __name__ == "__main__":
    build("/home/user/aitim/RU_battery_retailers.xlsx")
